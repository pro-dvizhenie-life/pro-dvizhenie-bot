"""Утилиты экспорта заявок в текстовые форматы."""

from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from io import BytesIO
from typing import Iterable, Iterator, List, Sequence

from django.core.exceptions import ImproperlyConfigured
from django.db.models import QuerySet
from django.http import HttpResponse, StreamingHttpResponse
from django.utils import timezone
from django.utils.encoding import smart_str

from ..models import Application, Question
from .form_runtime import build_answer_dict

__all__ = [
    "ApplicationExportDataset",
    "build_export_dataset",
    "export_applications_csv",
    "export_applications_xlsx",
]


@dataclass(slots=True)
class ApplicationExportDataset:
    """Представление таблицы экспорта заявок."""

    headers: List[str]
    rows: Iterator[List[str]]


class _Echo:
    """Фейковый буфер для StreamingHttpResponse."""

    def write(self, value):  # type: ignore[override]
        return value


def _collect_question_codes(queryset: QuerySet[Application]) -> List[str]:
    survey_ids = list(queryset.values_list("survey_id", flat=True).distinct())
    if not survey_ids:
        return []
    codes = (
        Question.objects.filter(step__survey_id__in=survey_ids)
        .values_list("code", flat=True)
        .distinct()
    )
    return sorted(code for code in codes if code)


def _format_datetime(value):
    if value is None:
        return ""
    aware = timezone.localtime(value) if timezone.is_aware(value) else value
    return aware.isoformat(timespec="seconds")


def _format_answer_value(value) -> str:
    if value in (None, "", [], {}):
        return ""
    if isinstance(value, bool):
        return "Да" if value else "Нет"
    if isinstance(value, (list, tuple, set)):
        rendered = ", ".join(_format_answer_value(item) for item in value)
        return rendered
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return smart_str(value)


# Используем явный chunk_size, чтобы поддержать QuerySet.iterator() после prefetch_related().
EXPORT_ITERATOR_CHUNK_SIZE = 512


def _build_rows(queryset: QuerySet[Application], question_codes: Sequence[str]) -> Iterator[List[str]]:
    for application in queryset.iterator(chunk_size=EXPORT_ITERATOR_CHUNK_SIZE):
        answers = build_answer_dict(application)
        row = [
            str(application.public_id),
            application.survey.code if application.survey else "",
            application.get_status_display(),
            application.status,
            application.applicant_type or "",
            str(application.current_stage),
            _format_datetime(application.created_at),
            _format_datetime(application.submitted_at),
            _format_datetime(application.updated_at),
        ]
        row.extend(_format_answer_value(answers.get(code)) for code in question_codes)
        yield row


def build_export_dataset(queryset: QuerySet[Application]) -> ApplicationExportDataset:
    question_codes = _collect_question_codes(queryset)
    headers = [
        "public_id",
        "survey_code",
        "status_label",
        "status",
        "applicant_type",
        "current_stage",
        "created_at",
        "submitted_at",
        "updated_at",
        *question_codes,
    ]
    rows = _build_rows(queryset, question_codes)
    return ApplicationExportDataset(headers=headers, rows=rows)


def export_applications_csv(queryset: QuerySet[Application], *, filename: str) -> StreamingHttpResponse:
    dataset = build_export_dataset(queryset)
    pseudo_buffer = _Echo()
    writer = csv.writer(pseudo_buffer)

    def stream() -> Iterable[str]:
        yield writer.writerow(dataset.headers)
        for row in dataset.rows:
            yield writer.writerow([smart_str(item) for item in row])

    response = StreamingHttpResponse(stream(), content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{filename}.csv"'
    return response


def export_applications_xlsx(queryset: QuerySet[Application], *, filename: str) -> HttpResponse:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - зависит от окружения
        raise ImproperlyConfigured("Для экспорта в XLSX требуется установить пакет 'openpyxl'.") from exc

    dataset = build_export_dataset(queryset)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Заявки"

    worksheet.append(dataset.headers)
    column_widths = [len(header) + 2 for header in dataset.headers]

    for row in dataset.rows:
        worksheet.append(row)
        for idx, value in enumerate(row, start=1):
            length = len(str(value)) + 2
            if length > column_widths[idx - 1]:
                column_widths[idx - 1] = min(length, 60)

    for idx, width in enumerate(column_widths, start=1):
        worksheet.column_dimensions[get_column_letter(idx)].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
    return response
